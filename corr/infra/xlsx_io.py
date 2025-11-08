from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..domain.types import Grid
from ..shared.utils import try_parse_float, fmt_serial
from ..shared.constants import TEXT


THIN = Border(
left=Side(style="thin", color="000000"),
right=Side(style="thin", color="000000"),
top=Side(style="thin", color="000000"),
bottom=Side(style="thin", color="000000"),
)


def load_xlsx_matrix(path: str) -> list[list[str]]:
wb = load_workbook(path, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = []
max_used = 0
for row in ws.iter_rows(values_only=True):
line = ["" if v is None else str(v) for v in row]
# подрезаем правые пустые
last = len(line)-1
while last >= 0 and (str(line[last]).strip()==""):
last -= 1
used = last+1
rows.append(line[:used])
max_used = max(max_used, used)
# выравниваем по ширине
out = [ (r + [""]*(max_used - len(r))) for r in rows ]
return out




def save_xlsx(path: str, grid: Grid, bg_rgb_func):
wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
rows = len(grid.rows); cols = grid.cols
for r in range(rows):
for c in range(cols):
txt = grid.rows[r].cells[c].text
f = try_parse_float(txt)
val = int(round(f)) if (f is not None and abs(f - int(round(f)))<1e-9 and c==0) else (f if f is not None else txt)
cell = ws.cell(row=r+1, column=c+1, value=val)
cell.alignment = Alignment(horizontal="center", vertical="center")
bg_rgb, fg_rgb = bg_rgb_func(r, c)
cell.fill = PatternFill(fill_type="solid", start_color=bg_rgb, end_color=bg_rgb)
cell.font = Font(name="Arial", size=11, color=fg_rgb)
cell.border = THIN
for c in range(1, cols+1):
ws.column_dimensions[get_column_letter(c)].width = 12
wb.save(path)