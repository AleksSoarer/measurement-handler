import sys
import re
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QSpinBox, QPushButton, QTableWidget, QTableWidgetItem, QFileDialog,
    QMessageBox, QAbstractItemView, QFrame
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor

# xlsx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ODS
from odf.opendocument import OpenDocumentSpreadsheet, load
from odf.style import Style, TableCellProperties, TextProperties
from odf.table import Table, TableRow, TableCell
from odf.text import P

import os, tempfile, html
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtGui import QPainter, QPixmap, QImage, QTextDocument, QFont
from PyQt5.QtCore import QRect, QRectF, QSizeF, Qt
from PyQt5.QtWidgets import QTableView, QListView, QScrollArea
from pypdf import PdfReader, PdfWriter

from os.path import basename

# ---- Colors ----
GREEN = QColor("#C6EFCE")  # ok data
RED   = QColor("#FFC7CE")  # bad data
BLUE  = QColor("#9DC3E6")  # good data
WHITE = QColor("#FFFFFF")  #
BLACK = QColor("#000000")  # NoMeasure
TEXT  = QColor("#000000")  #
YELLOW = QColor("#FFF2CC") # changed nominal


# ---- Layout sizes ----
HDR_PANEL_HEIGHT = 140
TOL_PANEL_HEIGHT = 70
INFO_COL_WIDTH   = 250  # ширина левого фиксированного столбца

# ---- Special rows (0-based in MAIN table) ----
MEASURE_INDEX_ROW = 2    # третья строка (0-based)
HEADER_ROWS    = [3, 4]  # «шапка», показываем сверху в отдельном виджете
NOMINAL_ROW    = 4       # 5-я строка: номинал (информативно)
TOL_ROW        = 5       # 6-я строка: допуск (редактируется в панели)
FIRST_DATA_ROW = 6       # данные с 7-й строки

MAX_CELLS = 300_000

# ---- Export font size (для ODS и PDF) ----
EXPORT_FONT_PT = 11.0   # меняй одно число: шрифт в сохраняемых файлах

# ---- UI font size (только виджетам на экране) ----
UI_FONT_PT = 10.0

PDF_ABOUT_TEXT = (
    "С учетом изменений допустимых величин отклонений указанных размеров произвести окончательную отбраковку деталей.<br/><br/>"
    "Достоверность измерений некоторых размеров не подтверждена. Указанные изменения никак не влияют на конструкционные свойства деталей, "
    "собираемость изделия в целом, его механические свойства, габаритные размеры, массу и т. д. Степень влияния на радиотехнические свойства минимальна, "
    "что будет подтверждено при проведении обязательного контроля всех радиотехнических характеристик всех изделий. "
    "По результатам измерений радиотехнических характеристик необходимо внести изменения в конструкторскую документацию."
)

# ---- Helpers ----
def _collect_defective_serials(self):
    """Вернёт список серийников (колонка 0) для строк, помеченных как брак."""
    bad = []
    rows = self.table.rowCount()
    for r in range(FIRST_DATA_ROW, rows):
        try:
            if self._is_row_defective(r):
                it = self.table.item(r, 0)
                sn = (it.text().strip() if it else "").strip()
                bad.append(sn or f"ROW {r}")
        except Exception:
            # На всякий — пропускаем проблемную строку, чтобы не уронить экспорт
            continue
    return bad

def _render_textpage_to_pdf(self, out_path: str, html_body: str):
    """Печатает одну текстовую страницу в PDF через QTextDocument."""
    printer = QPrinter(QPrinter.HighResolution)
    printer.setOutputFormat(QPrinter.PdfFormat)
    printer.setOutputFileName(out_path)
    try:
        # Новые Qt
        from PyQt5.QtGui import QPageLayout, QPageSize
        from PyQt5.QtCore import QMarginsF
        layout = QPageLayout(QPageSize(QPageSize.A4), QPageLayout.Portrait, QMarginsF(10,10,10,10))
        printer.setPageLayout(layout)
    except Exception:
        # Старые Qt
        printer.setPageSize(QPrinter.A4)
        printer.setOrientation(QPrinter.Portrait)

    doc = QTextDocument()
    doc.setDefaultFont(QFont("Arial", 11))
    doc.setHtml(html_body)
    doc.print_(printer)


def _fmt_serial(s: str) -> str:
    """Вернуть '283' вместо '283.0' (и '283,0'). Остальное — без изменений."""
    f = try_parse_float(s)
    if f is not None:
        i = int(round(f))
        if abs(f - i) < 1e-9:
            return str(i)
    return s or ""

def try_parse_float(s: str):
    if s is None:
        return None
    s = s.strip()
    if not s:
        return None
    up = s.upper()
    if up in ("Y", "N", "Z", "NM"):
        return None
    candidate = s.replace('−', '-').replace(',', '.')
    try:
        return float(candidate)
    except ValueError:
        return None

def _extract_text_from_cell(cell: TableCell) -> str:
    parts = []
    for p in cell.getElementsByType(P):
        for node in getattr(p, 'childNodes', []):
            data = getattr(node, 'data', None)
            if data:
                parts.append(str(data))
    text = "".join(parts).strip()
    if not text:
        v = cell.getAttribute('value')
        if v is not None:
            text = str(v)
    return text

def _cell_has_content(cell: TableCell) -> bool:
    v = cell.getAttribute('value')
    if v not in (None, ""):
        return True
    for p in cell.getElementsByType(P):
        for node in getattr(p, 'childNodes', []):
            if getattr(node, 'data', None):
                return True
    return False


class MiniOdsEditor(QWidget):
    def __init__(self):
        super().__init__()
        self._tol_cache = []
        self._in_cell_style = False
        self.setWindowTitle("Контроль допусков")
        self.resize(1280, 840)

        root = QVBoxLayout(self)

        # --- Controls ---
        ctrl = QHBoxLayout()

        self.sb_cols = QSpinBox(); self.sb_cols.setRange(1, 2000); self.sb_cols.setValue(8)

        self.sb_rows = QSpinBox(); self.sb_rows.setRange(1, 5000); self.sb_rows.setValue(12)

        self.btn_open = QPushButton("Открыть .ods"); self.btn_open.clicked.connect(self.open_ods)
        ctrl.addWidget(self.btn_open)

        self.btn_open_xlsx = QPushButton("Открыть .xlsx"); self.btn_open_xlsx.clicked.connect(self.open_xlsx)
        ctrl.addWidget(self.btn_open_xlsx)

        self.btn_save = QPushButton("Сохранить в .ods"); self.btn_save.clicked.connect(self.save_to_ods)
        ctrl.addWidget(self.btn_save)

        self.btn_save_xlsx = QPushButton("Сохранить в .xlsx"); self.btn_save_xlsx.clicked.connect(self.save_to_xlsx)
        ctrl.addWidget(self.btn_save_xlsx)

        self.btn_export_merged = QPushButton("PDF: таблица → чертёж → брак")
        self.btn_export_merged.setToolTip("Склеить: таблица (1-й лист), затем выбранный чертёж, затем лист 'Брак'")
        self.btn_export_merged.clicked.connect(self.export_report_pdf)  # <-- новое имя!
        ctrl.addWidget(self.btn_export_merged)

        ctrl.addStretch()
        root.addLayout(ctrl)

        self._nonnumeric_tol_cols = set()  # столбцы с «D9/6H…» — не анализируем
        self._slash_tol = {}

        # ======= TOP PANELS =======
        # left fixed header info (rows 1–2 of col0)
        self.info_header_table = QTableWidget(len(HEADER_ROWS), 1, self)
        self._setup_left_fixed_table(self.info_header_table, HDR_PANEL_HEIGHT)

        # center header (rows 1–2, all columns)
        self.header_table = QTableWidget(len(HEADER_ROWS), 0, self)
        self._setup_top_table(self.header_table, HDR_PANEL_HEIGHT, font_inc=1.5)
        self.header_table.cellChanged.connect(self.on_hdr_cell_changed)

        # left fixed tolerance info (row 5 of col0)
        self.info_tol_table = QTableWidget(1, 1, self)
        self._setup_left_fixed_table(self.info_tol_table, TOL_PANEL_HEIGHT)

        # center tolerance (row 5, all columns)
        self.tolerance_table = QTableWidget(1, 0, self)
        self._setup_top_table(self.tolerance_table, TOL_PANEL_HEIGHT, font_inc=2.0)
        self.tolerance_table.cellChanged.connect(self.on_tol_cell_changed)


        # left main info (col0 for all rows, except hidden 1,2,5 — мы их тоже скрываем здесь)
        self.info_main_table = QTableWidget(0, 1, self)
        self._setup_left_main_table(self.info_main_table)

        # Заголовок для левого фикс-столбца (залипает) — СОЗДАЁМ ДО добавления в layout
        self.info_main_caption = QTableWidget(1, 1, self)
        self.info_main_caption.verticalHeader().setVisible(False)
        self.info_main_caption.horizontalHeader().setVisible(False)
        self.info_main_caption.setFixedHeight(32)
        self.info_main_caption.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.info_main_caption.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.info_main_caption.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.info_main_caption.setColumnWidth(0, INFO_COL_WIDTH)  # ширина сразу здесь
        cap_item = QTableWidgetItem("Серийный/измерение")
        cap_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.info_main_caption.setItem(0, 0, cap_item)
        self.info_main_caption.setFrameStyle(QFrame.NoFrame)
        self.info_main_caption.setShowGrid(False)

        self.info_main_caption.setStyleSheet(
            "QTableWidget::item { background: white; font-weight: 600; padding: 6px; }"
        )

        # RIGHT side: vertical stack of header + tolerance + main
        #right_stack = QVBoxLayout()
        # Для левой части создаём такую же «стек»-колонку
        #left_stack = QVBoxLayout()

        # нет промежутков и полей — полосы «прилипают»
        #for lay in (left_stack, right_stack):
        #    lay.setSpacing(0)
        #    lay.setContentsMargins(0, 0, 0, 0)

        # подпись для нижней строки ("Не в допуске")
        self.info_oos_caption = QTableWidget(1, 1, self)
        self.info_oos_caption.verticalHeader().setVisible(False)
        self.info_oos_caption.horizontalHeader().setVisible(False)
        self.info_oos_caption.setFixedHeight(32)
        self.info_oos_caption.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.info_oos_caption.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.info_oos_caption.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.info_oos_caption.setColumnWidth(0, INFO_COL_WIDTH)
        oos_cap_item = QTableWidgetItem("Не в допуске")
        oos_cap_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.info_oos_caption.setItem(0, 0, oos_cap_item)
        self.info_oos_caption.setFrameStyle(QFrame.NoFrame)
        self.info_oos_caption.setShowGrid(False)
        self.info_oos_caption.setStyleSheet(
            "QTableWidget::item { background: white; font-weight: 600; padding: 6px; }"
        )

        # полоса счётчиков "Не в допуске" под основной таблицей
        self.oos_table = QTableWidget(1, 0, self)
        self._setup_top_table(self.oos_table, height=34, font_inc=0.0)
        self.oos_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.oos_table.setFrameStyle(QFrame.NoFrame)
        self.oos_table.setShowGrid(False)
        self.oos_table.setStyleSheet("QTableWidget::item { font-weight: 600; }")


        # ======= CENTER AREA: left fixed main info + right main =======
        left_stack = QVBoxLayout()
        right_stack = QVBoxLayout()
        for lay in (left_stack, right_stack):
            lay.setSpacing(0)
            lay.setContentsMargins(0, 0, 0, 0)

        # LEFT
        left_stack.addWidget(self.info_header_table)
        left_sep1 = QFrame(); left_sep1.setFrameShape(QFrame.HLine); left_sep1.setFrameShadow(QFrame.Sunken)
        left_stack.addWidget(left_sep1)
        left_stack.addWidget(self.info_tol_table)
        left_sep2 = QFrame(); left_sep2.setFrameShape(QFrame.HLine); left_sep2.setFrameShadow(QFrame.Sunken)
        left_stack.addWidget(left_sep2)
        left_stack.addWidget(self.info_main_caption)
        left_stack.addWidget(self.info_main_table)
        left_stack.addWidget(self.info_oos_caption)

        # RIGHT
        right_stack.addWidget(self.header_table)
        right_sep1 = QFrame(); right_sep1.setFrameShape(QFrame.HLine); right_sep1.setFrameShadow(QFrame.Sunken)
        right_stack.addWidget(right_sep1)
        right_stack.addWidget(self.tolerance_table)
        self.order_table = QTableWidget(1, 0, self)
        self._setup_top_table(self.order_table, height=34, font_inc=0.0)
        self.order_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.order_table.setFrameStyle(QFrame.NoFrame)
        self.order_table.setShowGrid(False)
        self.order_table.setStyleSheet("QTableWidget::item { font-weight: 600; }")
        right_stack.addWidget(self.order_table)

        self.table = QTableWidget(0, 0, self)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setVisible(False)
        self.table.cellChanged.connect(self.on_cell_changed)
        right_stack.addWidget(self.table, 1)

        right_stack.addWidget(self.oos_table)

        # Оборачиваем стэки в виджеты
        left_container = QWidget(); left_container.setLayout(left_stack)
        right_container = QWidget(); right_container.setLayout(right_stack)

        # ЕДИНЫЙ печатаемый контейнер (его и будем рендерить в PDF)
        self.report_panel = QWidget()
        rep_layout = QHBoxLayout(self.report_panel)
        rep_layout.setContentsMargins(12, 12, 12, 12)
        rep_layout.setSpacing(12)
        rep_layout.addWidget(left_container)
        sep_mid = QFrame(); sep_mid.setFrameShape(QFrame.VLine); sep_mid.setFrameShadow(QFrame.Sunken)
        rep_layout.addWidget(sep_mid)
        rep_layout.addWidget(right_container, 1)

        # Добавляем в корневой лэйаут
        root.addWidget(self.report_panel, 1)

        f = self.table.font(); f.setPointSizeF(UI_FONT_PT); self.table.setFont(f)

        # левые подписи
        f = self.info_main_caption.font(); f.setPointSizeF(UI_FONT_PT); self.info_main_caption.setFont(f)
        f = self.info_oos_caption.font();  f.setPointSizeF(UI_FONT_PT); self.info_oos_caption.setFont(f)

        # левые таблицы (если вдруг не через _setup_* создавались)
        f = self.info_main_table.font(); f.setPointSizeF(UI_FONT_PT); self.info_main_table.setFont(f)
        f = self.info_header_table.font(); f.setPointSizeF(UI_FONT_PT); self.info_header_table.setFont(f)
        f = self.info_tol_table.font();    f.setPointSizeF(UI_FONT_PT); self.info_tol_table.setFont(f)

        # ==== TOTAL DEFECTS STRIP (отдельное окошко) ====
        total_bar = QHBoxLayout()
        total_bar.setContentsMargins(0, 6, 0, 0)
        total_bar.setSpacing(8)

        lbl_total = QLabel("Итого брак:")
        lbl_total.setStyleSheet("font-weight:600;")
        self.total_defects_lbl = QLabel("0")
        self.total_defects_lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.total_defects_lbl.setStyleSheet("font-weight:700; padding:4px 8px; border:1px solid #ccc; border-radius:6px;")

        total_bar.addStretch(1)
        total_bar.addWidget(lbl_total)
        total_bar.addWidget(self.total_defects_lbl)
        root.addLayout(total_bar)

        # --- Sync scrollbars & sizes ---
        # Horizontal: sync center header/tolerance with main
        self.table.horizontalScrollBar().valueChanged.connect(self.header_table.horizontalScrollBar().setValue)
        self.table.horizontalScrollBar().valueChanged.connect(self.tolerance_table.horizontalScrollBar().setValue)

        # Горизонтальный скролл: main <-> order_row
        self.table.horizontalScrollBar().valueChanged.connect(self.order_table.horizontalScrollBar().setValue)
        self.order_table.horizontalScrollBar().valueChanged.connect(self.table.horizontalScrollBar().setValue)

        # Vertical: sync left main info with main
        self.table.verticalScrollBar().valueChanged.connect(self.info_main_table.verticalScrollBar().setValue)
        self.info_main_table.verticalScrollBar().valueChanged.connect(self.table.verticalScrollBar().setValue)

        # Подстраиваем значения при смене диапазона (чтобы не «уплывал» offset)
        self.table.verticalScrollBar().rangeChanged.connect(
            lambda _min, _max: self.info_main_table.verticalScrollBar().setValue(
                self.table.verticalScrollBar().value()
            )
        )
        self.info_main_table.verticalScrollBar().rangeChanged.connect(
            lambda _min, _max: self.table.verticalScrollBar().setValue(
                self.info_main_table.verticalScrollBar().value()
            )
        )
        # Column width sync (center panels)
        self.table.horizontalHeader().sectionResized.connect(self._on_main_section_resized)
        # Row height sync (left main info <-> main)
        self.table.verticalHeader().sectionResized.connect(self._on_main_row_height_changed)

        # Edits in left fixed tables should update main col0
        self.info_header_table.cellChanged.connect(self.on_info_header_cell_changed)
        self.info_tol_table.cellChanged.connect(self.on_info_tol_cell_changed)
        self.info_main_table.cellChanged.connect(self.on_info_main_cell_changed)

        self._orig_tol_texts = []   # базовые значения допусков (строка TOL_ROW)
        self._changed_tols = {}     # {col: (old_text, new_text)}

        # Init
        self.build_table()

        # main <-> oos_table (горизонтально)
        self.table.horizontalScrollBar().valueChanged.connect(self.oos_table.horizontalScrollBar().setValue)
        # oos_table без собственных полос, но связь в обе стороны не помешает
        self.oos_table.horizontalScrollBar().valueChanged.connect(self.table.horizontalScrollBar().setValue)



    # ---------- UI setup helpers ----------
    def _apply_service_row_visibility(self):
        #Спрятать все служебные строки (до FIRST_DATA_ROW) из нижних таблиц.
        rows = self.table.rowCount()
        # в правой main-таблице
        for r in range(min(rows, FIRST_DATA_ROW)):
            self.table.setRowHidden(r, True)
        # в левой info_main — чтобы выравнивание не «плыло»
        for r in range(min(rows, FIRST_DATA_ROW)):
            if r < self.info_main_table.rowCount():
                self.info_main_table.setRowHidden(r, True)

    def _setup_top_table(self, tw: QTableWidget, height: int, font_inc: float = 0.0):
        tw.verticalHeader().setVisible(False)
        tw.horizontalHeader().setVisible(False)
        tw.setFixedHeight(height)
        rows = max(1, tw.rowCount())
        for r in range(rows):
            tw.setRowHeight(r, max(28, height // rows - 2))
        f = tw.font(); f.setPointSizeF(UI_FONT_PT + font_inc); tw.setFont(f)
        tw.setStyleSheet("QTableWidget::item { padding: 6px; }")
        tw.setEditTriggers(QAbstractItemView.AllEditTriggers)
        #tw.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        # вот эти две строки обеспечат отсутствие полос и колёсика
        tw.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        tw.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        tw.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)

    def _setup_left_fixed_table(self, tw: QTableWidget, height: int):
        tw.setColumnCount(1)
        tw.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        tw.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        tw.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        tw.verticalHeader().setVisible(False)
        tw.horizontalHeader().setVisible(False)
        tw.setEditTriggers(QAbstractItemView.AllEditTriggers)
        tw.setFixedHeight(height)
        for r in range(tw.rowCount()):
            tw.setRowHeight(r, max(28, height // max(1, tw.rowCount()) - 2))
        tw.setColumnWidth(0, INFO_COL_WIDTH)
        tw.setStyleSheet("QTableWidget::item { background: white; padding: 4px; }")
        f = tw.font(); f.setPointSizeF(UI_FONT_PT); tw.setFont(f)

    def _setup_left_main_table(self, tw: QTableWidget):
        tw.setColumnCount(1)
        tw.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        tw.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        tw.verticalHeader().setVisible(False)
        tw.horizontalHeader().setVisible(False)
        tw.setEditTriggers(QAbstractItemView.AllEditTriggers)
        tw.setColumnWidth(0, INFO_COL_WIDTH)
        #tw.setStyleSheet("QTableWidget::item { background: white; padding: 4px; }")
        tw.setStyleSheet("QTableWidget::item { padding: 4px; }")
        f = tw.font(); f.setPointSizeF(UI_FONT_PT); tw.setFont(f)




    #------- Хэлперы для контроля допусков -------

    # ==== Slash tolerance helpers (внутри MiniOdsEditor) ====
    @staticmethod
    def _tof(s: str) -> float:
        return float(str(s).strip().replace('−', '-').replace(',', '.'))

    def _parse_slash_tolerance(self, tol_str: str):
        """
        ' -0,025/-0,05 ' -> ( -0.05, -0.025 )  # по возрастанию
        """
        s = (tol_str or '').strip().replace(' ', '')
        m = re.fullmatch(fr'({self._NUM_RE})[\\/]({self._NUM_RE})', s)
        if not m:
            raise ValueError(f"Некорректный формат допуска через слеш: {tol_str!r}")
        d1, d2 = self._tof(m.group(1)), self._tof(m.group(2))
        return (d1, d2) if d1 <= d2 else (d2, d1)

    def _check_delta_with_slash_pair(self, delta: float, dev_pair):
        """
        delta — значение отклонения из ячейки (например, -0.12).
        dev_pair — (lo, hi) из 'a/b', отсортированы по возрастанию.
        """
        lo, hi = dev_pair
        return lo <= delta <= hi

    # совместимость со старыми вызовами
    def _check_value_with_slash_pair(self, nominal: float, value: float, dev_pair):
        return self._check_delta_with_slash_pair(value, dev_pair)

    # --- ОПП-хелперы (внутри класса MiniOdsEditor) ---
    _NUM_RE = r'[-−]?\d+(?:[.,]\d+)?'
    
    _OPP_RE = re.compile(
        r'^\s*([0-9]+(?:[.,][0-9]+)?)(?:\s*\(\s*ОПП\s*([0-9]+(?:[.,][0-9]+)?)\s*\))?\s*$',
        re.IGNORECASE
    )

    _OPP_SLASH_DECOR_RE = re.compile(
        r'^\s*(' + _NUM_RE + r'\s*[\\/]\s*' + _NUM_RE + r')\s*\(\s*ОПП\s*(' + _NUM_RE + r'\s*[\\/]\s*' + _NUM_RE + r')\s*\)\s*$',
        re.IGNORECASE
    )
    _NUM_DOT_NUM_RE = re.compile(r'^[0-9]+\.[0-9]+$')
    _INT_RE = re.compile(r'^[0-9]+$')

    _NUM_ONLY_RE   = re.compile(r'^\d+(?:[.,]\d+)?$', re.ASCII)  # 12 или 12.34 / 12,34
    _NUM_SLASH_RE  = re.compile(
        r'^\s*[-−]?\d+(?:[.,]\d+)?\s*[\\/]\s*[-−]?\d+(?:[.,]\d+)?\s*$'
    )


    # Токен калибра: буква+цифры ИЛИ цифры+буква; допускаем латиницу/кириллицу
    _FIT_TOKEN     = r'(?:[A-Za-zА-Яа-я]\d+|\d+[A-Za-zА-Яа-я])'
    _SYM_PAIR_RE   = re.compile(rf'^\s*{_FIT_TOKEN}(?:[ /]{_FIT_TOKEN})?\s*$', re.IGNORECASE)

    # Отображаемая нами декорация "old (ОПП new)" — чтобы уметь её распознать при редактировании
    _OPP_DECOR_RE  = re.compile(
        r'^\s*([0-9]+(?:[.,][0-9]+)?)\s*\(\s*ОПП\s*([0-9]+(?:[.,][0-9]+)?)\s*\)\s*$',
        re.IGNORECASE
    )

    # НОВОЕ: диапазон чисел через дефис/тире: "0.1-0.2", "0,1 – 0,2" и т.п.
    _NUM_RANGE_RE = re.compile(
        r'^\s*[0-9]+(?:[.,][0-9]+)?\s*[-–—]\s*[0-9]+(?:[.,][0-9]+)?\s*$'
    )

    def _extract_tol_kind_and_value(self, s: str):
        s = (s or '').strip()
        if not s:
            return ('empty', '')
        # numeric "old (ОПП new)"
        m = self._OPP_DECOR_RE.fullmatch(s)
        if m:
            return ('numeric', m.group(2))
        # slash "old (ОПП new)"
        m = self._OPP_SLASH_DECOR_RE.fullmatch(s)
        if m:
            return ('slash', m.group(2))
        # plain number
        if self._NUM_ONLY_RE.fullmatch(s):
            return ('numeric', s)
        # plain slash
        if self._NUM_SLASH_RE.fullmatch(s):
            return ('slash', s)
        # символика (D9/6H и проч.)
        if self._SYM_PAIR_RE.fullmatch(s):
            return ('symbolic', s)
        return ('invalid', s)

    def _is_numeric_or_decorated_tol(self, s: str) -> bool:
        s = (s or '').strip()
        return bool(self._NUM_ONLY_RE.fullmatch(s) or self._OPP_DECOR_RE.fullmatch(s))

    def _is_range_tol_text(self, s: str) -> bool:
        if not s:
            return False
        return bool(self._NUM_RANGE_RE.match(s.replace(',', '.')))

    def _looks_symbolic_tol(self, s: str) -> bool:
        """
        True, если это «символьный» допуск (буквы/слеш) или диапазон через дефис.
        Такие допуски не анализируем численно.
        """
        s = (s or "").strip()
        if not s:
            return False
        if self._is_range_tol_text(s):
            return True
        if "/" in s:
            return True
        if self._contains_letters(s):
            return True
        return False

    def _normalize_to_xdoty(self, s: str) -> str:
        s = (s or "").strip().replace(",", ".")
        if not s:
            return ""
        if self._INT_RE.fullmatch(s):
            return s + ".0"
        if self._NUM_DOT_NUM_RE.fullmatch(s):
            return s
        return ""  # невалидно как число

    def _is_numeric_tol_text(self, s: str) -> bool:
        return bool(self._NUM_ONLY_RE.fullmatch((s or '').strip()))
    
    def _contains_letters(self, s: str) -> bool:
        return any(ch.isalpha() for ch in (s or ""))

    def _tol_current_part(self, s: str) -> str:
        """Для расчётов: из 'old (ОПП new)' берём new; иначе число; возвращаем с точкой."""
        s = (s or "").strip()
        m = self._OPP_DECOR_RE.fullmatch(s)
        if m:
            return self._normalize_to_xdoty(m.group(2))  # нормализуем только ДЛЯ расчёта
        if self._NUM_ONLY_RE.fullmatch(s):
            return self._normalize_to_xdoty(s)
        return ""

    def _tol_base_left_part(self, s: str) -> str:
        """Левая часть (old) из нашей декорации; иначе число; возвращаем с точкой."""
        s = (s or "").strip()
        m = self._OPP_DECOR_RE.fullmatch(s)
        if m:
            return self._normalize_to_xdoty(m.group(1))
        if self._NUM_ONLY_RE.fullmatch(s):
            return self._normalize_to_xdoty(s)
        return ""

    def _canon_tol(self, s: str):
        val = self._normalize_to_xdoty((s or "").strip())
        if not val:
            return None
        try:
            return round(float(val), 9)
        except Exception:
            return None

    def _mark_tol_change(self, col: int):
        if col <= 0 or col >= self.table.columnCount():
            return

        cell = self.table.item(TOL_ROW, col)
        cur_raw = (cell.text() if cell else "").strip()
        base_raw = (self._orig_tol_texts[col] if col < len(self._orig_tol_texts) else "").strip()

        # если символика/диапазоны нечисловые (не слэш), выкидываем из учёта
        if col in self._nonnumeric_tol_cols and not self._is_slash_tol_text(cur_raw):
            self._changed_tols.pop(col, None)
            self._apply_tol_highlight()
            return

        # numeric
        cur_num = self._tol_current_part(cur_raw)
        base_num = self._tol_base_left_part(base_raw)

        # slash
        cur_sl = self._tol_current_slash_part(cur_raw)
        base_sl = self._slash_base_left_part(base_raw)

        changed = False
        if cur_sl or base_sl:
            c = self._canon_slash_pair(cur_sl or "")
            b = self._canon_slash_pair(base_sl or "")
            if c is None or b is None or c != b:
                changed = True
        else:
            c = self._canon_tol(cur_num)
            b = self._canon_tol(base_num)
            if (c is None) or (b is None) or (c != b):
                changed = True

        if changed:
            # в список изменений кладём «как показано» (чтобы в отчёте выглядело знакомо)
            new_disp = cur_sl if cur_sl else (cur_num or cur_raw)
            old_disp = base_sl if base_sl else (base_num or base_raw)
            self._changed_tols[col] = (old_disp, new_disp)
        else:
            self._changed_tols.pop(col, None)

        self._apply_tol_highlight()

    def _measure_label(self, c: int) -> str:
        """Подпись измерения для колонки c — из третьей строки, иначе номер колонки."""
        if 0 <= MEASURE_INDEX_ROW < self.table.rowCount():
            it = self.table.item(MEASURE_INDEX_ROW, c)
            lab = (it.text() if it else "").strip()
            if lab:
                return lab
        return str(c)

    def _apply_tol_highlight(self):
        """Заливка жёлтым тех «номеров измерений» в order_table, у которых допуски изменены."""
        cols = self.order_table.columnCount()
        self.order_table.blockSignals(True)
        for c in range(cols):
            it = self.order_table.item(0, c)
            if it is None:
                it = QTableWidgetItem("")
                self.order_table.setItem(0, c, it)
            if c == 0:
                it.setBackground(WHITE)
            else:
                it.setBackground(YELLOW if c in self._changed_tols else WHITE)
        self.order_table.blockSignals(False)

    def _changed_tolerances_html(self) -> str:
        if not self._changed_tols:
            return ""
        items = []
        for c in sorted(self._changed_tols.keys()):
            old_txt, new_txt = self._changed_tols[c]
            label = self._measure_label(c)
            if self._is_slash_tol_text(old_txt) or self._is_slash_tol_text(new_txt):
                phrase = (
                    f"Размер <b>{html.escape(label)}</b>: – принять диапазон допустимых отклонений "
                    f"<i>{html.escape(new_txt)}</i> мкм (вместо <i>{html.escape(old_txt)}</i> мкм)."
                )
            else:
                phrase = (
                    f"Размер <b>{html.escape(label)}</b>: – принять допустимую величину отклонения от номинального размера "
                    f"равной <i>{html.escape(new_txt)}</i> мкм (вместо <i>{html.escape(old_txt)}</i> мкм)."
                )
            items.append(f"<li>{phrase}</li>")
        return "<h3>Изменённые допуски:</h3><ul>" + "".join(items) + "</ul>"
    

    def _is_slash_tol_text(self, s: str) -> bool:
        return bool(self._NUM_SLASH_RE.fullmatch((s or '').strip()))

    def _canon_slash_pair(self, s: str):
        """Вернёт (lo, hi) как float или None, если не слэш/непарсится."""
        try:
            lo, hi = self._parse_slash_tolerance(s)
            return (round(float(lo), 9), round(float(hi), 9))
        except Exception:
            return None

    def _tol_current_slash_part(self, s: str) -> str:
        """
        Из 'old (ОПП new)' вернёт 'new', из 'a/b' вернёт 'a/b', иначе ''.
        """
        s = (s or '').strip()
        # old/new как пара через слэш
        opp_slash = re.fullmatch(
            rf'\s*({self._NUM_RE}\s*[\\/]\s*{self._NUM_RE})\s*\(\s*ОПП\s*({self._NUM_RE}\s*[\\/]\s*{self._NUM_RE})\s*\)\s*',
            s, re.IGNORECASE
        )
        if opp_slash:
            return opp_slash.group(2)
        if self._NUM_SLASH_RE.fullmatch(s):
            return s
        return ""

    def _slash_base_left_part(self, s: str) -> str:
        """Левая часть для слэша из 'old (ОПП new)' или сам 'a/b'."""
        s = (s or '').strip()
        opp_slash = re.fullmatch(
            rf'\s*({self._NUM_RE}\s*[\\/]\s*{self._NUM_RE})\s*\(\s*ОПП\s*({self._NUM_RE}\s*[\\/]\s*{self._NUM_RE})\s*\)\s*',
            s, re.IGNORECASE
        )
        if opp_slash:
            return opp_slash.group(1)
        if self._NUM_SLASH_RE.fullmatch(s):
            return s
        return ""
    def _count_total_and_good(self):
        rows = self.table.rowCount()
        total = 0
        good = 0
        for r in range(FIRST_DATA_ROW, rows):
            if not self._row_is_empty_measurements(r):
                total += 1
                if not self._is_row_defective(r):
                    good += 1
        return total, good

    # ---------- Coloring rules ----------
    def _qcolor_to_xlsx_rgb(self, qc: QColor) -> str:
        """
        Возвращает 'RRGGBB' для openpyxl. QColor.name() даёт '#RRGGBB' — срежем '#'.
        """
        try:
            return qc.name()[1:].upper()
        except Exception:
            return "FFFFFF"

    def _cell_fill_for_bg(self, qc: QColor) -> PatternFill:
        rgb = self._qcolor_to_xlsx_rgb(qc)
        return PatternFill(fill_type="solid", start_color=rgb, end_color=rgb)

    def _font_for_cell(self, fg: QColor) -> Font:
        # шрифт экспорта + цвет текста
        rgb = self._qcolor_to_xlsx_rgb(fg)
        return Font(name="Arial", size=EXPORT_FONT_PT, color=rgb)

    def _get_tol(self, col):
        if col <= 0:
            return None
        if 0 <= col < len(self._tol_cache):
            return self._tol_cache[col]
        return None

    def recolor_cell(self, it: QTableWidgetItem, row=None, col=None):
        if not it:
            return
        if self._in_cell_style:
            return
        self._in_cell_style = True
        try:
            if row is None or col is None:
                idx = self.table.indexFromItem(it)
                row = idx.row(); col = idx.column()

            # === NEW: ряды 0..3 всегда белые с чёрным текстом ===
            if 0 <= row <= 3:
                it.setBackground(WHITE)
                it.setForeground(TEXT)
                return

       

            text = (it.text() or "").strip()
            up = text.upper()

            # кол.0
            if col == 0:
                it.setBackground(WHITE); it.setForeground(TEXT); return

            # служебные
            if row in HEADER_ROWS or row in (NOMINAL_ROW, TOL_ROW):
                it.setBackground(WHITE); it.setForeground(TEXT); return

            if up == "NM":
                it.setBackground(BLACK); it.setForeground(WHITE); return
            if up in ("N", "Z"):
                it.setBackground(RED); it.setForeground(TEXT); return
            if up == "Y":
                it.setBackground(GREEN); it.setForeground(TEXT); return

            # числа и допуски
            f = try_parse_float(text)
            if (row >= FIRST_DATA_ROW) and (col > 0) and f is not None:
                # 1) слэш-допуск: в ячейке хранится Δ (отклонение), сверяем с диапазоном [lo, hi] без номинала
                pair = self._slash_tol.get(col)
                if pair is not None:
                    ok = self._check_delta_with_slash_pair(f, pair)
                    it.setBackground(BLUE if ok else RED)
                    it.setForeground(TEXT if it.background().color() != BLACK else WHITE)
                    return

                # 2) скалярный допуск (старое поведение: считаем, что в ячейке уже Δ)
                tol = self._get_tol(col)
                if tol is not None:
                    it.setBackground(BLUE if abs(f) <= tol else RED)
                    it.setForeground(TEXT if it.background().color() != BLACK else WHITE)
                    return

            # fallback без вызовов it.text() и хелперов
            if any(ch.isalpha() for ch in text):
                it.setBackground(RED);   it.setForeground(TEXT)
            elif any(ch.isdigit() for ch in text):
                it.setBackground(GREEN); it.setForeground(TEXT)
            else:
                it.setBackground(WHITE); it.setForeground(TEXT)
        finally:
            self._in_cell_style = False

    def recolor_all(self):
        try:
            self.table.blockSignals(True)
            for r in range(self.table.rowCount()):
                for c in range(self.table.columnCount()):
                    it = self.table.item(r, c)
                    if it:
                        self.recolor_cell(it, r, c)
        finally:
            self.table.blockSignals(False)

    def recheck_column(self, col: int):
        if col <= 0:
            return
        try:
            self.table.blockSignals(True)
            for r in range(FIRST_DATA_ROW, self.table.rowCount()):
                it = self.table.item(r, col)
                if it is None: continue
                self.recolor_cell(it, r, col)
        finally:
            self.table.blockSignals(False)

    def _rebuild_tol_cache(self):
        cols = self.table.columnCount()
        self._tol_cache = [None] * cols     # скалярные допуски (старое поведение)
        self._slash_tol = {}                # НОВОЕ: пары отклонений для слэша

        for c in range(cols):
            if c == 0:
                continue
            it = self.table.item(TOL_ROW, c)
            raw = ((it.text() if it else "") or "").strip()

            # если это слэш или "слэш с ОПП" — берём текущую часть
            part = self._tol_current_slash_part(raw)  # вернёт 'a/b' либо ''
            if part:
                try:
                    self._slash_tol[c] = self._parse_slash_tolerance(part)
                    self._nonnumeric_tol_cols.discard(c)
                except Exception:
                    self._nonnumeric_tol_cols.add(c)
                continue

            # 2) Символика/диапазоны/прочее — как раньше
            if c in self._nonnumeric_tol_cols:
                self._tol_cache[c] = None
                continue

            # 3) Чисто числовой или "old (ОПП new)" — старое поведение
            cur = self._tol_current_part(raw)   # берём new (из скобок), если есть
            self._tol_cache[c] = try_parse_float(cur) if cur else None

    
    def _is_row_defective(self, r: int) -> bool:
        """
        Строка бракована, если:
        - в любой ячейке c>=1 есть 'N' или 'Z' (а также 'Н'/'З'),
        - есть число вне допуска,
        - вся строка измерений пуста (для c>=1).
        'NM'/'НМ' и 'Y' не считаем браком.
        """
        cols = self.table.columnCount()
        if cols <= 1:
            return True  # нет измерений — считаем браком

        has_any_value = False

        for c in range(1, cols):
            it = self.table.item(r, c)
            txt = (it.text() if it else "").strip()
            if txt:
                has_any_value = True

            up = txt.upper()

            # буквенные маркеры
            if up in ("N", "Z", "Н", "З"):
                return True
            if up in ("NM", "НМ", "Y"):
                continue

            # числа -> проверяем по допуску
            f = try_parse_float(txt)
            if f is not None:
                pair = self._slash_tol.get(c)
                if pair is not None:
                    # дельта уже в ячейке; номинал не нужен
                    if not self._check_delta_with_slash_pair(f, pair):
                        return True
                else:
                    tol = self._get_tol(c)
                    if tol is not None and abs(f) > tol:
                        return True

        # пустая строка измерений — брак
        return not has_any_value
    
    # ==== default name helpers ====
    def _default_basename(self) -> str:
        """
        Берём имя текущего файла без расширения.
        Если в конце есть _revXY (XY = буквы/цифры), отрезаем.
        """
        path = getattr(self, "current_file_path", "") or ""
        base = os.path.splitext(os.path.basename(path))[0] if path else ""
        if base:
            base = re.sub(r'(?i)(_rev[0-9a-z]+)$', '', base)  # срежем только хвостовой _revXY
        return base or "table"

    def _suggest_save_path(self, ext: str, fallback: str) -> str:
        """
        Собираем полный путь для диалога сохранения:
        папка = папка исходника (если есть), имя = _default_basename + ext.
        """
        directory = os.path.dirname(getattr(self, "current_file_path", "") or "")
        name = f"{self._default_basename()}{ext}"
        full = os.path.join(directory, name) if directory else name
        return full or fallback

    def _print_widget_to_single_pdf(self, pdf_path, widget):
        """Печатает ЛЮБОЙ контейнер (со всеми дочерними) в один лист PDF.
        Перед печатью временно разворачивает вложенные таблицы/списки,
        чтобы в PDF попал весь контент без полос прокрутки.
        """
        if widget is None:
            raise RuntimeError("Передан пустой виджет для печати")

        # 1) временно «распрямим» содержимое
        _restore = self._expand_children_for_print(widget)

        try:
            widget.adjustSize()
            QApplication.processEvents()

            content_w = max(1, widget.width())
            content_h = max(1, widget.height())

            printer = QPrinter(QPrinter.HighResolution)
            printer.setResolution(300)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(pdf_path)
            printer.setPaperSize(QPrinter.A4)
            printer.setFullPage(True)
            printer.setOrientation(QPrinter.Landscape if content_w >= content_h else QPrinter.Portrait)

            painter = QPainter(printer)
            if not painter.isActive():
                raise RuntimeError("Не удалось активировать QPainter для печати PDF")

            # Целевая область страницы (в пикселях устройства)
            target = printer.pageRect(QPrinter.DevicePixel)

            # Масштаб с сохранением пропорций + центрирование
            sx = target.width() / float(content_w)
            sy = target.height() / float(content_h)
            scale = min(sx, sy)

            view_w = int(content_w * scale)
            view_h = int(content_h * scale)
            offset_x = int((target.width() - view_w) / 2)
            offset_y = int((target.height() - view_h) / 2)

            painter.save()
            painter.translate(offset_x, offset_y)
            painter.scale(scale, scale)
            widget.render(painter, flags=QWidget.DrawChildren)
            painter.restore()
            painter.end()
        finally:
            # 2) вернуть всё как было
            try:
                _restore()
            except Exception:
                pass

    def export_report_pdf(self):
        """
        НОВЫЙ порядок:
        1) таблица (одним листом),
        2) внешний PDF (чертёж), если задан,
        3) лист с информацией по браку и изменённым допускам.
        """
        QMessageBox.information(self, "Экспорт", "Запущен экспорт: Таблица → Чертёж → Брак/Допуски")

        # 0) чертёж (опционален)
        start_dir = os.path.dirname(getattr(self, "current_file_path", "") or "")
        in_path, _ = QFileDialog.getOpenFileName(self, "Выбери чертёж (PDF) — можно пропустить",
                                                start_dir, "PDF Files (*.pdf)")
        out_default = self._suggest_save_path(".pdf", "report.pdf")
        out_path, _ = QFileDialog.getSaveFileName(self, "Сохранить как...", out_default, "PDF Files (*.pdf)")

        if not out_path:
            return

        # 1) offscreen-копия таблицы
        flat = self._build_offscreen_table_for_pdf()
        fd1, tmp_tbl_pdf = tempfile.mkstemp(suffix=".pdf"); os.close(fd1)
        fd2, tmp_bad_pdf = tempfile.mkstemp(suffix=".pdf"); os.close(fd2)

        try:
            # крупный шрифт для экспорта
            try:
                f = flat.font(); f.setPointSizeF(EXPORT_FONT_PT); flat.setFont(f)
            except Exception:
                pass
            flat.resizeColumnsToContents()
            flat.resizeRowsToContents()

            # печать таблицы → tmp_tbl_pdf
            self._print_table_to_single_pdf(tmp_tbl_pdf, flat)

            # формируем страницу «Брак/Допуски»
            bad_sns = _collect_defective_serials(self)
            bad_html = ", ".join(html.escape(x) for x in bad_sns) if bad_sns else "—"
            total_bad = len(bad_sns)

            fname = os.path.basename(getattr(self, "current_file_path", "") or "")
            header = f"<p style='font-size:12pt;'><b>{html.escape(fname)}</b></p>" if fname else ""
            changed_block = self._changed_tolerances_html()
            total_parts, good_parts = self._count_total_and_good()

            text_page_html = (
                header +
                "<h2>Брак:</h2>"
                f"<p>{bad_html}</p>"
                f"<p><b>Всего деталей:</b> {total_parts}; "
                f"<b>Годных:</b> {good_parts}; "
                f"<b>Итого брак:</b> {total_bad}</p>"
                + (changed_block or "") +
                "<p><br/></p><p><br/></p>"
                f"<p>{PDF_ABOUT_TEXT}</p>"
            )
            _render_textpage_to_pdf(self, tmp_bad_pdf, text_page_html)

            # Склейка в порядке: ТАБЛИЦА -> ЧЕРТЁЖ (если есть) -> БРАК/ДОПУСКИ
            writer = PdfWriter()

            # Таблица
            r_tbl = PdfReader(tmp_tbl_pdf)
            for p in r_tbl.pages:
                writer.add_page(p)

            # Чертёж (все страницы)
            if in_path:
                try:
                    r_in = PdfReader(in_path)
                    if getattr(r_in, "is_encrypted", False):
                        try:
                            r_in.decrypt("")
                        except Exception:
                            QMessageBox.warning(self, "Чертёж пропущен", "Выбранный PDF зашифрован, пропускаю чертёж.")
                            r_in = None
                    if r_in:
                        for p in r_in.pages:
                            writer.add_page(p)
                except Exception as e:
                    QMessageBox.warning(self, "Чертёж пропущен", f"Не удалось прочитать чертёж:\n{e}")

            # Брак/Допуски
            r_bad = PdfReader(tmp_bad_pdf)
            for p in r_bad.pages:
                writer.add_page(p)

            with open(out_path, "wb") as f:
                writer.write(f)

            QMessageBox.information(self, "Готово", f"PDF сохранён:\n{out_path}")

        except Exception as e:
            QMessageBox.critical(self, "Провал", f"Не удалось собрать PDF:\n{e}")
        finally:
            for tmp in (tmp_tbl_pdf, tmp_bad_pdf):
                try: os.remove(tmp)
                except Exception: pass
            try: flat.deleteLater()
            except Exception: pass

    def _build_offscreen_table_for_pdf(self) -> QTableWidget:
        """Полная автономная копия ВСЕЙ таблицы (включая кол.0 и служебные строки),
        с экспортным кеглем и автоподбором размеров, не зависящая от UI."""
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        tw = QTableWidget(rows, cols)          # parent=None — вне UI
        tw.verticalHeader().setVisible(False)
        tw.horizontalHeader().setVisible(False)
        tw.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tw.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        tw.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        tw.setWordWrap(False)

        # экспортный шрифт
        f = tw.font()
        f.setPointSizeF(float(EXPORT_FONT_PT))
        tw.setFont(f)

        # контент + стили
        for r in range(rows):
            for c in range(cols):
                src = self.table.item(r, c)
                txt = src.text() if src else ""
                it = QTableWidgetItem(txt)
                if src:
                    it.setBackground(src.background())
                    it.setForeground(src.foreground())
                    it.setTextAlignment(src.textAlignment())
                else:
                    it.setTextAlignment(Qt.AlignCenter)
                tw.setItem(r, c, it)

        # автоподбор под новый шрифт
        tw.resizeColumnsToContents()
        tw.resizeRowsToContents()

        # у 0-й колонки держим минимальную ширину инфо-колонки
        w0 = max(INFO_COL_WIDTH, tw.sizeHintForColumn(0))
        tw.setColumnWidth(0, w0)

        # на случай, если sizeHint где-то 0
        for c in range(cols):
            if tw.columnWidth(c) <= 0:
                tw.setColumnWidth(c, max(1, tw.sizeHintForColumn(c)))
        for r in range(rows):
            if tw.rowHeight(r) <= 0:
                tw.setRowHeight(r, max(1, tw.sizeHintForRow(r)))

        QApplication.processEvents()
        return tw

    def _print_table_to_single_pdf(self, pdf_path, table: QTableWidget):
        """
        Печать QTableWidget на ОДИН лист с масштабированием по большей стороне.
        Печатаем ВСЕ содержимое (без полос прокрутки).
        """
        vh = table.verticalHeader()
        hh = table.horizontalHeader()
        fw = table.frameWidth() * 2

        content_w = int(fw + vh.width() + sum(table.columnWidth(c) for c in range(table.columnCount())))
        content_h = int(fw + hh.height() + sum(table.rowHeight(r) for r in range(table.rowCount())))
        if content_w <= 0 or content_h <= 0:
            raise RuntimeError("Таблица пуста — печатать нечего.")

        printer = QPrinter(QPrinter.HighResolution)
        printer.setResolution(300)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(pdf_path)
        printer.setPaperSize(QPrinter.A4)
        printer.setFullPage(True)
        printer.setOrientation(QPrinter.Landscape if content_w >= content_h else QPrinter.Portrait)

        painter = QPainter(printer)
        if not painter.isActive():
            raise RuntimeError("Не удалось активировать QPainter для печати PDF")

        target = printer.pageRect(QPrinter.DevicePixel)

        old_size = table.size()
        old_hpol = table.horizontalScrollBarPolicy()
        old_vpol = table.verticalScrollBarPolicy()

        try:
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            table.resize(content_w, content_h)
            QApplication.processEvents()

            sx = target.width() / float(content_w)
            sy = target.height() / float(content_h)
            scale = min(sx, sy)

            view_w = max(1, int(content_w * scale))
            view_h = max(1, int(content_h * scale))
            offset_x = int((target.width() - view_w) / 2)
            offset_y = int((target.height() - view_h) / 2)

            painter.setViewport(QRect(offset_x, offset_y, view_w, view_h))
            painter.setWindow(QRect(0, 0, int(content_w), int(content_h)))

            table.render(painter, flags=QWidget.DrawChildren)
        finally:
            painter.end()
            table.resize(old_size)
            table.setHorizontalScrollBarPolicy(old_hpol)
            table.setVerticalScrollBarPolicy(old_vpol)

    def _expand_children_for_print(self, root_widget):
        """Убираем скроллы и растягиваем виджеты, чтобы в PDF попал весь контент."""
        saved = []

        # Таблицы (QTableWidget/QTableView)
        for tv in root_widget.findChildren(QTableView):
            st = {
                "w": tv,
                "size": tv.size(),
                "hpol": tv.horizontalScrollBarPolicy(),
                "vpol": tv.verticalScrollBarPolicy(),
            }
            saved.append(st)
            try:
                tv.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                tv.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                tv.resizeColumnsToContents()
                tv.resizeRowsToContents()

                vh = tv.verticalHeader(); hh = tv.horizontalHeader()
                fw = tv.frameWidth() * 2
                total_w = fw + vh.width() + sum(tv.columnWidth(c) for c in range(tv.model().columnCount()))
                total_h = fw + hh.height() + sum(tv.rowHeight(r) for r in range(tv.model().rowCount()))
                tv.resize(max(1, total_w), max(1, total_h))
            except Exception:
                pass

        # Списки
        for lv in root_widget.findChildren(QListView):
            st = {
                "w": lv,
                "size": lv.size(),
                "hpol": lv.horizontalScrollBarPolicy(),
                "vpol": lv.verticalScrollBarPolicy(),
            }
            saved.append(st)
            try:
                lv.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                lv.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                m = lv.model()
                rows = m.rowCount() if m is not None else 0
                base_h = lv.sizeHintForRow(0) if rows > 0 else lv.fontMetrics().height() + 6
                total_h = lv.frameWidth()*2 + base_h*max(1, rows)
                total_w = max(lv.width(), lv.viewport().sizeHint().width() + lv.frameWidth()*2)
                lv.resize(max(1, total_w), max(1, total_h))
            except Exception:
                pass

        # Скролл-области
        for sa in root_widget.findChildren(QScrollArea):
            inner = sa.widget()
            st = {
                "w": sa,
                "size": sa.size(),
                "hpol": sa.horizontalScrollBarPolicy(),
                "vpol": sa.verticalScrollBarPolicy(),
                "inner_size": inner.size() if inner else None,
            }
            saved.append(st)
            try:
                sa.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                sa.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                if inner:
                    inner.adjustSize()
                    inner.resize(inner.sizeHint())
                    sa.resize(inner.width() + sa.frameWidth()*2,
                            inner.height() + sa.frameWidth()*2)
            except Exception:
                pass

        def _restore():
            for st in saved:
                w = st["w"]
                try:
                    if hasattr(w, "setHorizontalScrollBarPolicy"):
                        w.setHorizontalScrollBarPolicy(st.get("hpol", Qt.ScrollBarAsNeeded))
                    if hasattr(w, "setVerticalScrollBarPolicy"):
                        w.setVerticalScrollBarPolicy(st.get("vpol", Qt.ScrollBarAsNeeded))
                    w.resize(st["size"])
                    if isinstance(w, QScrollArea) and st.get("inner_size") and w.widget():
                        w.widget().resize(st["inner_size"])
                except Exception:
                    pass

        return _restore


    
    def _row_is_empty_measurements(self, r: int) -> bool:
        """True, если во всех ячейках c>=1 пусто (игнорируем служебные строки)."""
        cols = self.table.columnCount()
        if cols <= 1 or r < FIRST_DATA_ROW:
            return True
        for c in range(1, cols):
            it = self.table.item(r, c)
            if it and (it.text() or "").strip() != "":
                return False
        return True


    def _recompute_total_defects(self):
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        if rows <= FIRST_DATA_ROW:
            self.total_defects_lbl.setText("0")
            return

        total_bad = 0
        self._in_cell_style = True
        try:
            for r in range(FIRST_DATA_ROW, rows):
                is_bad = self._is_row_defective(r)
                if is_bad:
                    total_bad += 1

                it_left = self.info_main_table.item(r, 0)
                if it_left is not None:
                    it_left.setBackground(RED if is_bad else WHITE)
                    it_left.setForeground(WHITE if is_bad else TEXT)

                it0 = self.table.item(r, 0)
                if it0 is not None:
                    it0.setBackground(RED if is_bad else WHITE)
                    it0.setForeground(TEXT)

                # заливка пустой бракованной строки
                is_empty_line = self._row_is_empty_measurements(r)
                if is_bad and is_empty_line:
                    for c in range(1, cols):
                        itc = self.table.item(r, c) or QTableWidgetItem("")
                        if self.table.item(r, c) is None:
                            self.table.setItem(r, c, itc)
                        itc.setBackground(RED)
                        itc.setForeground(TEXT)
                else:
                    for c in range(1, cols):
                        itc = self.table.item(r, c)
                        if itc and (itc.text() or "").strip() == "":
                            itc.setBackground(WHITE)
                            itc.setForeground(TEXT)
        finally:
            self._in_cell_style = False

        self.total_defects_lbl.setText(str(total_bad))

    # ---------- Panels/Info sync helpers ----------
    
    def _ensure_panel_cols(self):
        cols = self.table.columnCount()

        # center header/tol panels
        for tw in (self.header_table, self.tolerance_table):
            if tw.columnCount() != cols:
                tw.blockSignals(True)
                tw.setColumnCount(cols)
                # init cells
                for r in range(tw.rowCount()):
                    for c in range(cols):
                        it = tw.item(r, c)
                        if it is None:
                            tw.setItem(r, c, QTableWidgetItem(""))
                        tw.item(r, c).setTextAlignment(Qt.AlignCenter)
                tw.blockSignals(False)
            # widths
            for c in range(cols):
                tw.setColumnWidth(c, self.table.columnWidth(c))
            # hide col0 – его отображают left-виджеты
            if cols > 0:
                tw.setColumnHidden(0, True)

        # left main info rows = точно как в main
        rows = self.table.rowCount()
        if self.info_main_table.rowCount() != rows:
            self.info_main_table.blockSignals(True)
            self.info_main_table.setRowCount(rows)
            for r in range(rows):
                it = self.info_main_table.item(r, 0)
                if it is None:
                    self.info_main_table.setItem(r, 0, QTableWidgetItem(""))
                self.info_main_table.item(r, 0).setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.info_main_table.blockSignals(False)
        # высоты строк и скрытие 1,2,5 для точного совпадения
        for r in range(rows):
            self.info_main_table.setRowHeight(r, self.table.rowHeight(r))
        #for r in HEADER_ROWS + [TOL_ROW]:
        #    if r < rows:
        #        self.info_main_table.setRowHidden(r, True)

        # высоты строк
        for r in range(rows):
            self.info_main_table.setRowHeight(r, self.table.rowHeight(r))
        # скрываем все служебные (0..FIRST_DATA_ROW-1)
        for r in range(min(rows, FIRST_DATA_ROW)):
            self.info_main_table.setRowHidden(r, True)
        
        # left header & tol fixed tables – ширина колонки уже задана, высоты держим = панелям
        # ничего дополнительно делать не нужно здесь
        self._sync_order_row()

        # --- синхронизировать нижнюю полосу (oos_table) с main
        cols = self.table.columnCount()
        if self.oos_table.columnCount() != cols:
            self.oos_table.blockSignals(True)
            self.oos_table.setColumnCount(cols)
            for c in range(cols):
                it = self.oos_table.item(0, c)
                if it is None:
                    it = QTableWidgetItem("")
                    self.oos_table.setItem(0, c, it)
                self.oos_table.item(0, c).setTextAlignment(Qt.AlignCenter)
            self.oos_table.blockSignals(False)

        # ширины колонок = как у main
        for c in range(cols):
            self.oos_table.setColumnWidth(c, self.table.columnWidth(c))

        # скрываем 0-й столбец (описательный)
        if cols > 0:
            self.oos_table.setColumnHidden(0, True)

    def _sync_header_from_main(self):
        self.header_table.blockSignals(True)
        cols = self.table.columnCount()
        for i, r_main in enumerate(HEADER_ROWS):
            for c in range(cols):
                src = self.table.item(r_main, c)
                txt = src.text() if src else ""
                it = self.header_table.item(i, c)
                if it is None:
                    it = QTableWidgetItem(""); self.header_table.setItem(i, c, it)
                it.setTextAlignment(Qt.AlignCenter)
                it.setText(txt); it.setBackground(WHITE)
        self.header_table.blockSignals(False)

        # left header info (col0)
        self.info_header_table.blockSignals(True)
        for i, r_main in enumerate(HEADER_ROWS):
            src = self.table.item(r_main, 0)
            txt = src.text() if src else ""
            it = self.info_header_table.item(i, 0)
            if it is None:
                it = QTableWidgetItem(""); self.info_header_table.setItem(i, 0, it)
            it.setText(txt)
        self.info_header_table.blockSignals(False)

    def _sync_order_row(self):
        """Подписи в верхней полосе (order_table) = номера измерений из третьей строки.
        Подсветка жёлтым — для измерений с изменённым допуском.
        """
        cols = self.table.columnCount()

        if self.order_table.columnCount() != cols:
            self.order_table.blockSignals(True)
            self.order_table.setColumnCount(cols)
            for c in range(cols):
                it = self.order_table.item(0, c)
                if it is None:
                    it = QTableWidgetItem("")
                    self.order_table.setItem(0, c, it)
                self.order_table.item(0, c).setTextAlignment(Qt.AlignCenter)
            self.order_table.blockSignals(False)

        for c in range(cols):
            self.order_table.setColumnWidth(c, self.table.columnWidth(c))

        if cols > 0:
            self.order_table.setColumnHidden(0, True)

        self.order_table.blockSignals(True)
        for c in range(cols):
            it = self.order_table.item(0, c)
            if it is None:
                it = QTableWidgetItem("")
                self.order_table.setItem(0, c, it)

            if c == 0:
                it.setText("")
                it.setBackground(WHITE)
            else:
                it.setText(self._measure_label(c))
                # важное: НЕ затираем подсветку; выставляем в соответствии с изменениями
                it.setBackground(YELLOW if c in self._changed_tols else WHITE)
        self.order_table.blockSignals(False)

        self._sync_bars_and_captions_height()
        self._apply_tol_highlight()

    def _snapshot_orig_tolerances(self):
        cols = self.table.columnCount()
        self._orig_tol_texts = []
        self._nonnumeric_tol_cols.clear()

        for c in range(cols):
            it = self.table.item(TOL_ROW, c)
            raw = (it.text().strip() if it else "")

            # 1) Слэш: plain 'a/b' ИЛИ 'a/b (ОПП x/y)' → берём ЛЕВУЮ часть как базу
            sl_base = self._slash_base_left_part(raw)
            if sl_base:
                self._orig_tol_texts.append(sl_base)
                display = raw
                self._nonnumeric_tol_cols.discard(c)
            # 2) Число: 'n' или 'n (ОПП m)' → берём левую часть как базу
            elif self._is_numeric_or_decorated_tol(raw):
                m = self._OPP_DECOR_RE.fullmatch(raw)
                old_disp = m.group(1) if m else raw
                self._orig_tol_texts.append(old_disp)
                display = raw
            # 3) Символика/диапазоны — в «нечисловые»
            else:
                self._nonnumeric_tol_cols.add(c)
                self._orig_tol_texts.append(raw)
                display = raw

            if it:
                it.setText(display)
            top = self.tolerance_table.item(0, c)
            if top:
                top.setText(display)

        self._changed_tols.clear()
        self._apply_tol_highlight()
        self._rebuild_tol_cache()

    
    def on_hdr_cell_changed(self, row, col):
        """
        Редактирование ячеек верхней шапки (header_table).
        Прокидываем текст в скрытые строки основной таблицы (HEADER_ROWS)
        и не запускаем перекраску (эти строки всегда белые).
        """
        if self._in_cell_style:
            return
        if row < 0 or col < 0:
            return

        main_row = HEADER_ROWS[row]

        src_item = self.header_table.item(row, col)
        txt = src_item.text() if src_item else ""

        it = self.table.item(main_row, col)
        if it is None:
            it = QTableWidgetItem("")
            self.table.setItem(main_row, col, it)

        try:
            self.table.blockSignals(True)
            it.setText(txt)
            it.setBackground(WHITE)  # служебные строки не красим
        finally:
            self.table.blockSignals(False)

        # Если редактировали 0-й столбец шапки — обновим левую фикс-таблицу для шапки
        if col == 0 and row < self.info_header_table.rowCount():
            try:
                self.info_header_table.blockSignals(True)
                left_it = self.info_header_table.item(row, 0)
                if left_it is None:
                    left_it = QTableWidgetItem("")
                    self.info_header_table.setItem(row, 0, left_it)
                left_it.setText(txt)
            finally:
                self.info_header_table.blockSignals(False)

    def _sync_tol_from_main(self):
        self.tolerance_table.blockSignals(True)
        cols = self.table.columnCount()
        for c in range(cols):
            src = self.table.item(TOL_ROW, c)
            txt = src.text() if src else ""
            it = self.tolerance_table.item(0, c)
            if it is None:
                it = QTableWidgetItem(""); self.tolerance_table.setItem(0, c, it)
            it.setTextAlignment(Qt.AlignCenter)
            it.setText(txt); it.setBackground(WHITE)
        self.tolerance_table.blockSignals(False)

        # left tol info (col0)
        self.info_tol_table.blockSignals(True)
        src = self.table.item(TOL_ROW, 0)
        txt = src.text() if src else ""
        it = self.info_tol_table.item(0, 0)
        if it is None:
            it = QTableWidgetItem(""); self.info_tol_table.setItem(0, 0, it)
        it.setText(txt)
        self.info_tol_table.blockSignals(False)

    def _sync_info_main_from_main(self):
        self.info_main_table.blockSignals(True)
        rows = self.table.rowCount()
        for r in range(rows):
            src = self.table.item(r, 0)
            txt = src.text() if src else ""
            it = self.info_main_table.item(r, 0)
            if it is None:
                it = QTableWidgetItem("")
                self.info_main_table.setItem(r, 0, it)
            it.setText(_fmt_serial(txt))  # показываем целые
        self.info_main_table.blockSignals(False)

    def _on_main_section_resized(self, logicalIndex, oldSize, newSize):
        if logicalIndex < self.tolerance_table.columnCount():
            self.tolerance_table.setColumnWidth(logicalIndex, newSize)
        if logicalIndex < self.header_table.columnCount():
            self.header_table.setColumnWidth(logicalIndex, newSize)
        if logicalIndex < self.order_table.columnCount():
            self.order_table.setColumnWidth(logicalIndex, newSize)
        if logicalIndex < self.oos_table.columnCount():
            self.oos_table.setColumnWidth(logicalIndex, newSize)

    def _on_main_row_height_changed(self, logicalIndex, oldSize, newSize):
        if 0 <= logicalIndex < self.info_main_table.rowCount():
            self.info_main_table.setRowHeight(logicalIndex, newSize)
        self._sync_bars_and_captions_height()

    def on_info_header_cell_changed(self, row, col):
        main_row = HEADER_ROWS[row]
        txt = self.info_header_table.item(row, col).text() if self.info_header_table.item(row, col) else ""
        it = self.table.item(main_row, 0)
        if it is None:
            it = QTableWidgetItem(""); self.table.setItem(main_row, 0, it)
        try:
            self.table.blockSignals(True)
            it.setText(txt); it.setBackground(WHITE)
        finally:
            self.table.blockSignals(False)

    def on_info_tol_cell_changed(self, row, col):
        txt = self.info_tol_table.item(row, col).text() if self.info_tol_table.item(row, col) else ""
        it = self.table.item(TOL_ROW, 0)
        if it is None:
            it = QTableWidgetItem(""); self.table.setItem(TOL_ROW, 0, it)
        try:
            self.table.blockSignals(True)
            it.setText(txt); it.setBackground(WHITE)
        finally:
            self.table.blockSignals(False)

    def on_info_main_cell_changed(self, row, col):
        raw = self.info_main_table.item(row, col).text() if self.info_main_table.item(row, col) else ""
        txt = _fmt_serial(raw)
        self.info_main_table.item(row, 0).setText(txt)
        it = self.table.item(row, 0)
        # обновляем левый виджет (на случай, если пользователь ввёл 283.0)
        self.info_main_table.item(row, 0).setText(txt)
        if it is None:
            it = QTableWidgetItem(""); self.table.setItem(row, 0, it)
        try:
            self.table.blockSignals(True)
            it.setText(txt) 
            #it.setBackground(WHITE)
        finally:
            self.table.blockSignals(False)

    def _sync_order_and_caption_height(self):
        """Высота полосы нумерации и левой подписи = высоте первой рабочей строки."""
        # высота первой рабочей строки (после служебных)
        if self.table.rowCount() > FIRST_DATA_ROW:
            h = self.table.rowHeight(FIRST_DATA_ROW)
        else:
            h = max(28, self.order_table.rowHeight(0))

        # полоса нумерации
        self.order_table.setRowHeight(0, h)
        self.order_table.setFixedHeight(h)  # без рамок достаточно ровно h

        # левая подпись «Серийный номер»
        self.info_main_caption.setFixedHeight(h)

    def _sync_bars_and_captions_height(self):
        """Высота полос (order/oos) и левых подписей = высоте первой рабочей строки."""
        if self.table.rowCount() > FIRST_DATA_ROW:
            h = self.table.rowHeight(FIRST_DATA_ROW)
        else:
            h = 34  # запасной

        # верхняя полоса нумерации
        self.order_table.setRowHeight(0, h)
        self.order_table.setFixedHeight(h)
        # нижняя полоса счётчиков
        self.oos_table.setRowHeight(0, h)
        self.oos_table.setFixedHeight(h)

        # левые подписи
        self.info_main_caption.setFixedHeight(h)
        self.info_oos_caption.setFixedHeight(h)

    # ---------- UI callbacks ----------
    def build_table(self):
        cols = max(self.sb_cols.value(), 1)
        rows = max(self.sb_rows.value(), FIRST_DATA_ROW + 1)
        try:
            self.table.blockSignals(True)
            self.table.setColumnCount(cols)
            self.table.setRowCount(rows)
            self.table.clearContents()

            for r in range(rows):
                for c in range(cols):
                    it = self.table.item(r, c)
                    if it is None:
                        it = QTableWidgetItem(""); self.table.setItem(r, c, it)
                    it.setTextAlignment(Qt.AlignCenter)
                    it.setBackground(WHITE)

    
            # скрываем колонку 0 в main — её показывают левые таблицы
            if cols > 0:
                self.table.setColumnHidden(0, True)
            self._apply_service_row_visibility()

            # выровнять панели/левые таблицы и залить данными
            self._ensure_panel_cols()
            self._sync_header_from_main()
            self._sync_tol_from_main()
            self._rebuild_tol_cache()
            self._sync_info_main_from_main()
            self._sync_order_row()
            self._recompute_oos_counts()
            self._sync_bars_and_captions_height()
            self.recolor_all()
            self._recompute_total_defects()

        finally:
            self.table.blockSignals(False)
            self.table.horizontalScrollBar().setValue(0)
            self.order_table.horizontalScrollBar().setValue(0)
            self._snapshot_orig_tolerances()

    def _format_tol_with_opp_display(self, new_display_str: str, col: int) -> str:
        """
        Для чисел и для слэша. Если new == old по ЧИСЛАМ, вернём просто old.
        """
        old_disp = self._orig_tol_texts[col] if col < len(self._orig_tol_texts) else ""
        new_disp = (new_display_str or "").strip()

        # случай: слэш ↔ слэш
        if self._is_slash_tol_text(old_disp) and self._is_slash_tol_text(new_disp):
            old_pair = self._canon_slash_pair(old_disp)
            new_pair = self._canon_slash_pair(new_disp)
            if old_pair and new_pair and old_pair == new_pair:
                return old_disp or new_disp
            return f"{old_disp} (ОПП {new_disp})" if old_disp else new_disp

        # случай: число ↔ число
        old_num = self._canon_tol(old_disp)
        new_num = self._canon_tol(new_disp)
        if old_num is not None and new_num is not None and abs(old_num - new_num) < 1e-12:
            return old_disp or new_disp

        # разный тип (число vs слэш) — просто показываем ОПП
        return f"{old_disp} (ОПП {new_disp})" if old_disp else new_disp

    def _format_tol_with_opp(self, cur_txt: str, col: int) -> str:
        """Показываем: 'old (ОПП new)'; если old==new — только 'old'."""
        new_val = self._normalize_to_xdoty(cur_txt)
        old_val = self._normalize_to_xdoty(self._orig_tol_texts[col] if col < len(self._orig_tol_texts) else "")
        if not new_val and not old_val:
            return ""
        if not old_val:
            return new_val
        # если равны — пишем только одно
        if self._canon_tol(new_val) is not None and self._canon_tol(new_val) == self._canon_tol(old_val):
            return old_val
        return f"{old_val} (ОПП {new_val})"

    def on_tol_cell_changed(self, row, col):
        if col < 0:
            return

        raw_in = self.tolerance_table.item(0, col).text() if self.tolerance_table.item(0, col) else ""
        txt = (raw_in or "").strip()
        prev = self.table.item(TOL_ROW, col).text() if self.table.item(TOL_ROW, col) else ""

        kind, val_disp = self._extract_tol_kind_and_value(txt)

        # пусто -> откат
        if kind == 'empty':
            it_top = self.tolerance_table.item(0, col) or QTableWidgetItem("")
            if self.tolerance_table.item(0, col) is None:
                self.tolerance_table.setItem(0, col, it_top)
            it_top.setText(prev)
            return

        # мусор -> откат + предупреждение
        if kind == 'invalid':
            it_top = self.tolerance_table.item(0, col) or QTableWidgetItem("")
            if self.tolerance_table.item(0, col) is None:
                self.tolerance_table.setItem(0, col, it_top)
            it_top.setText(prev)
            QApplication.beep()
            QMessageBox.warning(
                self, "Неверный формат допуска",
                "Разрешены только:\n"
                "• число (с точкой или запятой), например: 0.15 или 0,15\n"
                "• два числа через слэш: 0.10/0.25 или 0,10/0,25\n"
                "• калибры: D6, 6H, D6 7H, D9/6H"
            )
            return

        # символика и 'a/b' — не анализируем численно
        if kind == 'slash':
            new_disp = val_disp.strip()  # сюда уже прилетает "новая" часть thanks to _extract_tol_kind_and_value
            # валидация пары
            try:
                _ = self._parse_slash_tolerance(new_disp)
            except Exception:
                it_top = self.tolerance_table.item(0, col) or QTableWidgetItem("")
                if self.tolerance_table.item(0, col) is None:
                    self.tolerance_table.setItem(0, col, it_top)
                it_top.setText(prev)
                QApplication.beep()
                QMessageBox.warning(self, "Неверный формат допуска", "Ожидалось два числа через слэш, например: -0,025/-0,05")
                return

            self._nonnumeric_tol_cols.discard(col)

            old_disp = self._orig_tol_texts[col] if col < len(self._orig_tol_texts) else ""
            if not old_disp:
                # впервые задали слэш — фиксируем базу
                if col >= len(self._orig_tol_texts):
                    self._orig_tol_texts.extend([""] * (col + 1 - len(self._orig_tol_texts)))
                self._orig_tol_texts[col] = new_disp
                display = new_disp
            else:
                # если по числам равно — показываем просто old
                if self._canon_slash_pair(old_disp) == self._canon_slash_pair(new_disp):
                    display = old_disp
                else:
                    display = f"{old_disp} (ОПП {new_disp})"

            # в таблицы (низ и верх)
            it = self.table.item(TOL_ROW, col) or QTableWidgetItem("")
            if self.table.item(TOL_ROW, col) is None:
                self.table.setItem(TOL_ROW, col, it)
            try:
                self.table.blockSignals(True)
                it.setText(display); it.setBackground(WHITE)
            finally:
                self.table.blockSignals(False)

            it_top = self.tolerance_table.item(0, col) or QTableWidgetItem("")
            if self.tolerance_table.item(0, col) is None:
                self.tolerance_table.setItem(0, col, it_top)
            it_top.setText(display)

            # обновить кэш и метрики
            self._rebuild_tol_cache()
            self.recheck_column(col)
            self._mark_tol_change(col)
            self._recompute_oos_counts()
            self._recompute_total_defects()
            return

        # numeric — включаем автодекор с сохранением ВИДА
        if kind == 'numeric':
            self._nonnumeric_tol_cols.discard(col)

            old_disp = self._orig_tol_texts[col] if col < len(self._orig_tol_texts) else ""
            new_disp = val_disp  # как ввёл пользователь (с точкой ИЛИ запятой)

            if not old_disp:
                # впервые задают допуск в колонке: фиксируем как базовый "old"
                if col >= len(self._orig_tol_texts):
                    self._orig_tol_texts.extend([""] * (col + 1 - len(self._orig_tol_texts)))
                self._orig_tol_texts[col] = new_disp
                display = new_disp
            else:
                display = self._format_tol_with_opp_display(new_disp, col)

            it = self.table.item(TOL_ROW, col) or QTableWidgetItem("")
            if self.table.item(TOL_ROW, col) is None:
                self.table.setItem(TOL_ROW, col, it)
            try:
                self.table.blockSignals(True)
                it.setText(display); it.setBackground(WHITE)
            finally:
                self.table.blockSignals(False)

            it_top = self.tolerance_table.item(0, col) or QTableWidgetItem("")
            if self.tolerance_table.item(0, col) is None:
                self.tolerance_table.setItem(0, col, it_top)
            it_top.setText(display)

            self._rebuild_tol_cache()
            self.recheck_column(col)
            self._mark_tol_change(col)
            self._recompute_oos_counts()
            self._recompute_total_defects()

    def on_cell_changed(self, row, col):
        # отражаем возможные изменения скрытых строк в панели/левых таблицах
        if row in HEADER_ROWS:
            self._sync_header_from_main()
            return
        if row == TOL_ROW:
            self._sync_tol_from_main()
            self._rebuild_tol_cache()
            self.recheck_column(col)
            self._mark_tol_change(col)
            return
        if col == 0:
            # изменили скрытую кол.0 в main (через код/загрузку)
            self._sync_info_main_from_main()
            return

        # обычные данные — раскрасить
        it = self.table.item(row, col)
        self.recolor_cell(it, row, col)

        if col > 0 and row >= FIRST_DATA_ROW:
            self._recompute_oos_counts()
            self._recompute_total_defects()
        elif col == 0 and row >= FIRST_DATA_ROW:
            # меняли серийник — могло стать "пустая строка" по факту? да, но считаем по c>=1.
            self._recompute_total_defects()

    # не в допуске 
    def _recompute_oos_counts(self):
        """Посчитать количество ячеек вне допуска по каждому столбцу (c >= 1).
        Вне допуска:
        - 'N'/'Z' (и 'Н'/'З'),
        - числа, у которых abs(value) > tol (если tol задан).
        'Y' и 'NM'/'НМ' игнорируем.
        """
        cols = self.table.columnCount()
        rows = self.table.rowCount()
        if cols == 0 or rows == 0:
            return

        self._ensure_panel_cols()

        self.oos_table.blockSignals(True)
        try:
            for c in range(cols):
                cell = self.oos_table.item(0, c)
                if cell is None:
                    cell = QTableWidgetItem("")
                    self.oos_table.setItem(0, c, cell)

                if c == 0:
                    cell.setText("")
                    cell.setBackground(WHITE)
                    continue

                tol = self._get_tol(c)
                cnt = 0

                pair = self._slash_tol.get(c)
                tol  = self._get_tol(c)

                for r in range(FIRST_DATA_ROW, rows):
                    it = self.table.item(r, c)
                    if not it:
                        continue
                    txt = (it.text() or "").strip()
                    if not txt:
                        continue

                    up = txt.upper()
                    if up in ("N", "Z", "Н", "З"):
                        cnt += 1
                        continue
                    if up in ("Y", "NM", "НМ"):
                        continue

                    f = try_parse_float(txt)
                    if f is None:
                        continue

                    if pair is not None:
                        if not self._check_delta_with_slash_pair(f, pair):
                            cnt += 1
                    elif tol is not None:
                        if abs(f) > tol:
                            cnt += 1
                    else:
                        # толеранс нечисловой — пропускаем
                        pass

                cell.setText(str(cnt))
                cell.setBackground(WHITE)
        finally:
            self.oos_table.blockSignals(False)

    # ---------- ODS I/O ----------
    def save_to_ods(self):
        default_name = self._suggest_save_path(".ods", "table.ods")
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить как…", default_name, "ODS (*.ods)")
        if not path: return

        self.current_file_path = path
        
        doc = OpenDocumentSpreadsheet()


        # Общие текстовые настройки для ячеек (шрифт)
        def _txt_props(color="#000000"):
            return TextProperties(fontsize=f"{EXPORT_FONT_PT}pt", color=color)

        # ЯВНЫЕ бордеры для каждой стороны. LibreOffice так надёжнее.
        BORDER_SPEC = "0.75pt solid #808080"

        def _cellprops(bg=None):
            kw = {
                "bordertop":    BORDER_SPEC,
                "borderbottom": BORDER_SPEC,
                "borderleft":   BORDER_SPEC,
                "borderright":  BORDER_SPEC,
            }
            if bg:
                kw["backgroundcolor"] = bg
            return TableCellProperties(**kw)

        # Стили ячеек: каждый со своими бордерами
        style_green = Style(name="cellGreen", family="table-cell")
        style_green.addElement(_cellprops("#C6EFCE"))
        style_green.addElement(_txt_props("#000000"))
        doc.automaticstyles.addElement(style_green)

        style_red = Style(name="cellRed", family="table-cell")
        style_red.addElement(_cellprops("#FFC7CE"))
        style_red.addElement(_txt_props("#000000"))
        doc.automaticstyles.addElement(style_red)

        style_blue = Style(name="cellBlue", family="table-cell")
        style_blue.addElement(_cellprops("#9DC3E6"))
        style_blue.addElement(_txt_props("#000000"))
        doc.automaticstyles.addElement(style_blue)

        style_white = Style(name="cellWhite", family="table-cell")
        style_white.addElement(_cellprops("#FFFFFF"))
        style_white.addElement(_txt_props("#000000"))
        doc.automaticstyles.addElement(style_white)

        # NM: чёрный фон + белый текст + те же бордеры
        style_black = Style(name="cellBlack", family="table-cell")
        style_black.addElement(_cellprops("#000000"))
        style_black.addElement(_txt_props("#FFFFFF"))
        doc.automaticstyles.addElement(style_black)

        t = Table(name="Sheet1"); doc.spreadsheet.addElement(t)
        
        rows = self.table.rowCount(); cols = self.table.columnCount()
        for r in range(rows):
            tr = TableRow(); t.addElement(tr)
            for c in range(cols):
                it = self.table.item(r, c)
                text = it.text() if it else ""
                bg = it.background().color() if it else WHITE
                hexbg = bg.name().upper() if hasattr(bg, "name") else "#FFFFFF"

                if   hexbg == "#C6EFCE": stylename = style_green
                elif hexbg == "#FFC7CE": stylename = style_red
                elif hexbg == "#9DC3E6": stylename = style_blue
                elif hexbg == "#000000": stylename = style_black
                else:                    stylename = style_white

                f = try_parse_float(text)
                if f is not None:
                    if c == 0 and abs(f - int(round(f))) < 1e-9:
                        ival = int(round(f))
                        cell = TableCell(valuetype="float", value=ival, stylename=stylename)
                        cell.addElement(P(text=str(ival)))
                    else:
                        cell = TableCell(valuetype="float", value=f, stylename=stylename)
                        cell.addElement(P(text=str(f)))
                else:
                    # для строки в первом столбце тоже подчистим внешний вид
                    txt = _fmt_serial(text) if c == 0 else text
                    cell = TableCell(valuetype="string", stylename=stylename)
                    cell.addElement(P(text=txt))
                tr.addElement(cell)

        doc.save(path)
        self.setWindowTitle(f"Контроль допусков. Имя открытого файла:   {basename(path)}")
        self.btn_save.setText("ODS Сохранено ✓")

    def open_ods(self):
        path, _ = QFileDialog.getOpenFileName(self, "Открыть…", "", "ODS (*.ods)")
        
        if not path:
            return
        
        self.current_file_path = path
        self.setWindowTitle(f"Контроль допусков. Имя открытого файла:   {basename(path)}")
        
        from PyQt5.QtWidgets import QApplication
        QApplication.setOverrideCursor(Qt.WaitCursor)

        try:
            doc = load(path)
            tables = doc.spreadsheet.getElementsByType(Table)
            if not tables:
                return
            sheet = tables[0]

            # ---------- 1) ПЕРВЫЙ ПРОХОД: считаем нужное кол-во строк/колонок ----------
            row_specs = []   # список: (useful_cols, rrep, blocks) ; blocks = [(txt, crep), ...]
            max_cols = 0
            last_content_row_idx = -1
            cum_rows = 0

            for row in sheet.getElementsByType(TableRow):
                rrep = int(row.getAttribute('numberrowsrepeated') or 1)

                blocks = []
                col_idx = 0
                last_non_empty = -1

                for cell in row.getElementsByType(TableCell):
                    crep = int(cell.getAttribute('numbercolumnsrepeated') or 1)
                    txt = _extract_text_from_cell(cell)
                    blocks.append((txt, crep))
                    if (txt or "").strip() != "":
                        last_non_empty = col_idx + crep - 1
                    col_idx += crep

                useful_cols = last_non_empty + 1  # 0 если строка полностью пустая
                row_specs.append((useful_cols, rrep, blocks))

                if useful_cols > 0:
                    last_content_row_idx = cum_rows + rrep - 1
                    if useful_cols > max_cols:
                        max_cols = useful_cols

                cum_rows += rrep

            # сколько строк реально нужно загрузить:
            needed_rows = max(last_content_row_idx + 1, FIRST_DATA_ROW + 1)
            if max_cols <= 0:
                # вообще нет контента — подготовим минимальную таблицу
                self.table.blockSignals(True); self.table.setUpdatesEnabled(False)
                try:
                    self.table.clearContents()
                    self.table.setRowCount(1); self.table.setColumnCount(1)
                    self.sb_rows.setValue(1); self.sb_cols.setValue(1)
                    it = QTableWidgetItem(""); it.setTextAlignment(Qt.AlignCenter); it.setBackground(WHITE)
                    self.table.setItem(0, 0, it)
                finally:
                    self.table.setUpdatesEnabled(True); self.table.blockSignals(False)
                return

            # ограничение по общему числу ячеек
            max_rows_by_cells = max(1, MAX_CELLS // max(1, max_cols))
            use_rows = min(needed_rows, max_rows_by_cells)
            truncated = use_rows < needed_rows
            use_cols = max_cols

            # ---------- 2) ВТОРОЙ ПРОХОД: разворачиваем строки до use_rows/use_cols ----------
            rows_buf = []
            for useful_cols, rrep, blocks in row_specs:
                if len(rows_buf) >= use_rows:
                    break

                # соберём строку до useful_cols (не разворачиваем хвостовую пустоту)
                useful = min(useful_cols, use_cols)
                row_line = []
                remain = useful
                for txt, crep in blocks:
                    if remain <= 0:
                        break
                    vis = crep if crep <= remain else remain
                    if vis > 0:
                        row_line.extend([txt] * vis)
                        remain -= vis

                # добьём до use_cols пустыми, чтобы ширина везде одинаковая
                if useful < use_cols:
                    row_line.extend([""] * (use_cols - useful))

                # добавим rrep раз, но не больше нужного
                times = min(rrep, use_rows - len(rows_buf))
                for _ in range(times):
                    rows_buf.append(list(row_line))

                # чтобы UI не «замирал» на больших файлах
                if (len(rows_buf) & 0xFF) == 0:
                    QApplication.processEvents()

            # ---------- 3) Загрузка в QTableWidget ----------
            try:
                self.table.blockSignals(True); self.table.setUpdatesEnabled(False)
                self.table.clearContents()

                final_rows = max(use_rows, FIRST_DATA_ROW + 1)
                self.table.setRowCount(final_rows)
                self.table.setColumnCount(use_cols)
                self.sb_rows.setValue(final_rows)
                self.sb_cols.setValue(use_cols)

                # фактические строки из файла
                for r in range(use_rows):
                    row_vals = rows_buf[r]
                    for c in range(use_cols):
                        txt = row_vals[c] if c < len(row_vals) else ""
                        it = self.table.item(r, c)
                        if it is None:
                            it = QTableWidgetItem("")
                            self.table.setItem(r, c, it)
                        it.setTextAlignment(Qt.AlignCenter)
                        it.setText(_fmt_serial(txt) if c == 0 else txt)

                # хвост до final_rows — чистые белые строки
                for r in range(use_rows, final_rows):
                    for c in range(use_cols):
                        it = self.table.item(r, c)
                        if it is None:
                            it = QTableWidgetItem("")
                            self.table.setItem(r, c, it)
                        it.setTextAlignment(Qt.AlignCenter)
                        it.setBackground(WHITE)
            finally:
                self.table.setUpdatesEnabled(True); self.table.blockSignals(False)

            # ---------- 4) Синхронизация/пересчёт ----------
            self._apply_service_row_visibility()
            if self.table.columnCount() > 0:
                self.table.setColumnHidden(0, True)

            self._ensure_panel_cols()
            self._sync_header_from_main()
            self._sync_tol_from_main()
            self._rebuild_tol_cache()          # ВАЖНО: после загрузки!
            self._sync_info_main_from_main()
            self._sync_order_row()
            self.table.horizontalScrollBar().setValue(0)
            self.order_table.horizontalScrollBar().setValue(0)
            self.recolor_all()
            self._recompute_oos_counts()       # теперь tol на месте
            self._sync_bars_and_captions_height()
            self._recompute_total_defects()
            self._snapshot_orig_tolerances()

            if truncated:
                QMessageBox.information(
                    self, "Файл урезан",
                    f"Загружено {use_rows}×{use_cols} (лимит ≈ {MAX_CELLS:,} ячеек)."
                )
        finally:
            QApplication.restoreOverrideCursor()


    def _print_whole_table_to_single_pdf(self, pdf_path, table, font_pt):
        """Печатает ИМЕННО всю основную таблицу (включая скрытые служебные строки и колонку 0)
        на один лист PDF с масштабированием, не затрагивая UI."""
        if table.rowCount() == 0 or table.columnCount() == 0:
            raise RuntimeError("Таблица пуста — печатать нечего.")

        # --- сохранить состояние, чтобы потом вернуть ---
        hidden_rows = [table.isRowHidden(r) for r in range(table.rowCount())]
        col0_hidden = table.isColumnHidden(0)
        old_font = table.font()
        old_hpol = table.horizontalScrollBarPolicy()
        old_vpol = table.verticalScrollBarPolicy()
        old_size = table.size()

        try:
            # Показать всё: колонку 0 и служебные строки
            table.setColumnHidden(0, False)
            for r in range(table.rowCount()):
                table.setRowHidden(r, False)

            # Временный шрифт только для печати
            f = old_font
            f.setPointSizeF(float(font_pt))
            table.setFont(f)

            # Подогнать размеры ячеек под шрифт
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            table.resizeColumnsToContents()
            table.resizeRowsToContents()
            table.clearSpans()  # на всякий

            # Полный размер содержимого
            vh = table.verticalHeader(); hh = table.horizontalHeader(); fw = table.frameWidth() * 2
            content_w = int(fw + vh.width() + sum(table.columnWidth(c) for c in range(table.columnCount())))
            content_h = int(fw + hh.height() + sum(table.rowHeight(r) for r in range(table.rowCount())))
            content_w = max(1, content_w)
            content_h = max(1, content_h)
            table.resize(content_w, content_h)
            QApplication.processEvents()

            # Подготовка принтера
            printer = QPrinter(QPrinter.HighResolution)
            printer.setResolution(300)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(pdf_path)
            printer.setPaperSize(QPrinter.A4)
            printer.setFullPage(True)
            printer.setOrientation(QPrinter.Landscape if content_w >= content_h else QPrinter.Portrait)

            painter = QPainter(printer)
            if not painter.isActive():
                raise RuntimeError("Не удалось активировать QPainter для печати PDF")

            # Прямоугольник страницы в пикселях устройства
            target = printer.pageRect(QPrinter.DevicePixel)

            # Масштаб с сохранением пропорций
            sx = target.width() / float(content_w)
            sy = target.height() / float(content_h)
            scale = min(sx, sy)

            view_w = max(1, int(content_w * scale))
            view_h = max(1, int(content_h * scale))
            offset_x = int((target.width() - view_w) / 2)
            offset_y = int((target.height() - view_h) / 2)

            # Вьюпорт/оконная система — всё целиком на страницу
            painter.setViewport(offset_x, offset_y, view_w, view_h)
            painter.setWindow(0, 0, content_w, content_h)
            table.render(painter, flags=QWidget.DrawChildren)
            painter.end()

        finally:
            # --- вернуть состояние ---
            table.setFont(old_font)
            table.setHorizontalScrollBarPolicy(old_hpol)
            table.setVerticalScrollBarPolicy(old_vpol)
            table.resize(old_size)
            table.setColumnHidden(0, col0_hidden)
            for r, was_hidden in enumerate(hidden_rows):
                table.setRowHidden(r, was_hidden)


    def open_xlsx(self):
        path, _ = QFileDialog.getOpenFileName(self, "Открыть…", "", "Excel (*.xlsx)")
        if not path:
            return

        self.current_file_path = path
        self.setWindowTitle(f"Контроль допусков. Имя открытого файла:   {basename(path)}")

        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            if not wb.sheetnames:
                QMessageBox.warning(self, "Пусто", "В книге нет листов.")
                return
            ws = wb[wb.sheetnames[0]]

            # Собираем буфер строк как str, аккуратно
            rows_buf = []
            max_used_cols = 0
            last_content_row_idx = -1

            # read_only worksheet -> iter_rows быстрый и не жрёт память как бегемот
            for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                # Приводим к строкам, но без ломания запятых/точек:
                # openpyxl уже дал числа как float/int → строку через str(), для колонки 0 применим _fmt_serial
                line = []
                last_non_empty = -1
                for cidx, v in enumerate(row, start=1):
                    if v is None:
                        s = ""
                    else:
                        if isinstance(v, float):
                            # Excel хранит как float → строка с точкой; нам норм, дальше раскраска через try_parse_float
                            s = str(v)
                        elif isinstance(v, int):
                            s = str(v)
                        else:
                            s = str(v)
                    if s.strip() != "":
                        last_non_empty = cidx - 1
                    line.append(s)

                useful_cols = last_non_empty + 1
                if useful_cols > 0:
                    last_content_row_idx = ridx - 1
                    max_used_cols = max(max_used_cols, useful_cols)
                    # обрежем «хвост» пустого справа
                    line = line[:useful_cols]
                    rows_buf.append(line)
                else:
                    rows_buf.append([])  # для согласования индексов

            if max_used_cols <= 0:
                # Вообще пусто — создадим минимальную таблицу
                self.table.blockSignals(True); self.table.setUpdatesEnabled(False)
                try:
                    self.table.clearContents()
                    self.table.setRowCount(1); self.table.setColumnCount(1)
                    self.sb_rows.setValue(1); self.sb_cols.setValue(1)
                    it = QTableWidgetItem(""); it.setTextAlignment(Qt.AlignCenter); it.setBackground(WHITE)
                    self.table.setItem(0, 0, it)
                finally:
                    self.table.setUpdatesEnabled(True); self.table.blockSignals(False)
                return

            needed_rows = max(last_content_row_idx + 1, FIRST_DATA_ROW + 1)
            max_rows_by_cells = max(1, MAX_CELLS // max(1, max_used_cols))
            use_rows = min(needed_rows, max_rows_by_cells)
            truncated = use_rows < needed_rows
            use_cols = max_used_cols

            # Загрузка в QTableWidget
            try:
                self.table.blockSignals(True); self.table.setUpdatesEnabled(False)
                self.table.clearContents()

                final_rows = max(use_rows, FIRST_DATA_ROW + 1)
                self.table.setRowCount(final_rows)
                self.table.setColumnCount(use_cols)
                self.sb_rows.setValue(final_rows)
                self.sb_cols.setValue(use_cols)

                for r in range(use_rows):
                    row_vals = rows_buf[r] if r < len(rows_buf) else []
                    for c in range(use_cols):
                        raw = row_vals[c] if c < len(row_vals) else ""
                        # кол.0 — серийник «283.0» → «283»
                        txt = _fmt_serial(raw) if c == 0 else raw
                        it = self.table.item(r, c)
                        if it is None:
                            it = QTableWidgetItem("")
                            self.table.setItem(r, c, it)
                        it.setTextAlignment(Qt.AlignCenter)
                        it.setText(txt)
                        it.setBackground(WHITE)
            finally:
                self.table.setUpdatesEnabled(True); self.table.blockSignals(False)

            # Синхронизация и пересчёты (как в open_ods)
            self._apply_service_row_visibility()
            if self.table.columnCount() > 0:
                self.table.setColumnHidden(0, True)

            self._ensure_panel_cols()
            self._sync_header_from_main()
            self._sync_tol_from_main()
            self._rebuild_tol_cache()
            self._sync_info_main_from_main()
            self._sync_order_row()
            self.table.horizontalScrollBar().setValue(0)
            self.order_table.horizontalScrollBar().setValue(0)
            self.recolor_all()
            self._recompute_oos_counts()
            self._sync_bars_and_captions_height()
            self._recompute_total_defects()
            self._snapshot_orig_tolerances()

            if truncated:
                QMessageBox.information(
                    self, "Файл урезан",
                    f"Загружено {use_rows}×{use_cols} (лимит ≈ {MAX_CELLS:,} ячеек)."
                )

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть XLSX:\n{e}")
        finally:
            QApplication.restoreOverrideCursor()

    def save_to_xlsx(self):
        default_name = self._suggest_save_path(".xlsx", "table.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить как…", default_name, "Excel (*.xlsx)")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        rows = self.table.rowCount()
        cols = self.table.columnCount()

        thin_black = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
)
        
        # Пишем значения и минимальную стилизацию: фон и цвет текста
        for r in range(rows):
            for c in range(cols):
                it = self.table.item(r, c)
                txt = (it.text() if it else "") or ""

                # Значение: пытаемся сохранить число числом; иначе строку
                f = try_parse_float(txt)
                if f is not None:
                    val = int(round(f)) if (abs(f - int(round(f))) < 1e-9) else float(f)
                else:
                    val = txt

                cell = ws.cell(row=r+1, column=c+1, value=val)

                # Выравнивание по центру, как в UI
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

                # Фон и шрифт
                bg = (it.background().color() if it else WHITE)
                fg = (it.foreground().color() if it else TEXT)

                if c == 0:
                    fg = TEXT

                cell.border = thin_black
                try:
                    cell.fill = self._cell_fill_for_bg(bg)
                except Exception:
                    pass
                try:
                    cell.font = self._font_for_cell(fg)
                except Exception:
                    pass

        # Немного ширины для читаемости
        for c in range(1, cols+1):
            ws.column_dimensions[get_column_letter(c)].width = max(10, min(50, self.table.columnWidth(c-1) // 7 or 12))

        # Высоты строк примерные
        #for r in range(1, rows+1):
        #    ws.row_dimensions[r].height = max(14, min(120, self.table.rowHeight(r-1) * 0.75))

        try:
            wb.save(path)
            self.current_file_path = path
            self.setWindowTitle(f"Контроль допусков. Имя открытого файла:   {basename(path)}")
            self.btn_save_xlsx.setText("XLSX Сохранено ✓")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить XLSX:\n{e}")

def main():
    app = QApplication(sys.argv)
    w = MiniOdsEditor()
    w.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
